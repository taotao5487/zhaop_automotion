import asyncio
import base64
import hashlib
import json
import logging
import mimetypes
import re
import shutil
import subprocess
import tempfile
import zipfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import unquote, urljoin, urlparse
from xml.etree import ElementTree as ET

import aiohttp
import yaml
from bs4 import BeautifulSoup, NavigableString, Tag

from crawler.core.crawler import (
    CrawlerConfig,
    _looks_like_safeline_js_challenge,
    _looks_like_upstream_error_page,
    _looks_like_yunsuo_verify_page,
)
from crawler.core.wechat_review_builder import WechatReviewBuilder

logger = logging.getLogger(__name__)

try:
    from readability import Document
except Exception:  # pragma: no cover - optional dependency
    Document = None

try:
    from markdownify import markdownify as html_to_markdown
except Exception:  # pragma: no cover - optional dependency
    html_to_markdown = None

try:
    import fitz
except Exception:  # pragma: no cover - optional dependency
    fitz = None

try:
    import mammoth
except Exception:  # pragma: no cover - optional dependency
    mammoth = None

try:
    from PIL import Image, ImageDraw, ImageFont
except Exception:  # pragma: no cover - optional dependency
    Image = None
    ImageDraw = None
    ImageFont = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

try:
    from scrapling.fetchers import DynamicFetcher, StealthyFetcher
except Exception:  # pragma: no cover - optional dependency
    DynamicFetcher = None
    StealthyFetcher = None

DEFAULT_DETAIL_RULES_PATH = "config/detail_rules.yaml"
DEFAULT_DETAIL_OUTPUT_ROOT = "data/exports/details"
DEFAULT_WECHAT_REVIEW_OUTPUT_ROOT = "data/exports/wechat_review"
DEFAULT_REMOVE_SELECTORS = [
    "script",
    "style",
    "nav",
    "footer",
    "aside",
    "iframe",
    "form",
    "noscript",
    ".breadcrumb",
    ".breadcrumbs",
    ".header",
    ".footer",
    ".sidebar",
    ".share",
    ".shares",
    ".related",
    ".recommend",
    ".recommended",
    ".advert",
    ".advertisement",
    ".ads",
    ".pagination",
    ".pager",
    ".copy",
    ".copyright",
    ".toolbar",
]
TRAILING_NOISE_PATTERNS = (
    r"^责任编辑[:：]?",
    r"^打印本页$",
    r"^关闭窗口$",
    r"^上一篇",
    r"^下一篇",
    r"^返回顶部$",
    r"^浏览次数[:：]?\d*$",
)
ATTACHMENT_EXTENSIONS = {
    ".pdf",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".zip",
    ".rar",
    ".7z",
    ".csv",
    ".txt",
}
RENDERABLE_ATTACHMENT_EXTENSIONS = {".pdf", ".doc", ".docx", ".xlsx"}
ATTACHMENT_HINT_KEYWORDS = ("附件", "报名表", "岗位表", "职位表", "应聘表", "下载")
INLINE_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}
POSITION_ATTACHMENT_KEYWORDS = ("岗位", "职位", "一览表", "需求表", "岗位表", "职位表", "招聘计划")
FORM_ATTACHMENT_KEYWORDS = ("报名表", "登记表", "申请表", "信息收集表", "信息表", "汇总表", "扫码表")
LONG_TEXT_COLUMN_KEYWORDS = (
    "条件",
    "职责",
    "内容",
    "备注",
    "说明",
    "经历",
    "资历",
    "资格",
    "范围",
)
SHORT_TEXT_COLUMN_KEYWORDS = (
    "人数",
    "年龄",
    "学历",
    "学位",
    "序号",
    "编号",
    "代码",
    "性别",
)
MEDIUM_TEXT_COLUMN_KEYWORDS = ("岗位", "科室", "名称", "部门", "方向", "专业", "要求")
RAW_ATTACHMENT_SELECTORS = (
    "#attachContainer a[href]",
    ".attach-container a[href]",
    ".attachBox a[href]",
    ".fk-attach-download-table a[href]",
    ".attachName-container a[href]",
    "a.attachName[href]",
)
LEGACY_DOC_STANDARD_HEADERS = [
    "序号",
    "单位",
    "科室",
    "岗位",
    "数量（人）",
    "学历",
    "专业",
    "执业（职称）条件",
    "其他条件",
]
LEGACY_DOC_PAGE_MARKER_PATTERN = re.compile(r"page\s+\\\*\s+mergeformat", re.I)


@dataclass
class DetailRule:
    content_selectors: List[str] = field(default_factory=list)
    remove_selectors: List[str] = field(default_factory=list)
    attachment_selectors: List[str] = field(default_factory=list)
    prefer_readability: bool = False
    force_browser_fetch: Optional[str] = None


class DetailPipeline:
    """公告详情增强流水线。"""

    def __init__(
        self,
        crawler: Any,
        detail_rules_path: str = DEFAULT_DETAIL_RULES_PATH,
        output_root: str = DEFAULT_DETAIL_OUTPUT_ROOT,
        review_output_root: str = DEFAULT_WECHAT_REVIEW_OUTPUT_ROOT,
        max_items: Optional[int] = None,
    ):
        self.crawler = crawler
        self.detail_rules_path = detail_rules_path
        self.output_root = Path(output_root)
        self.review_output_root = Path(review_output_root)
        self.max_items = max_items
        self.rules = self._load_rules(detail_rules_path)
        self.review_builder = WechatReviewBuilder(output_root=review_output_root)

    @staticmethod
    def _load_rules(detail_rules_path: str) -> Dict[str, DetailRule]:
        path = Path(detail_rules_path)
        if not path.exists():
            return {}

        try:
            payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        except Exception as exc:
            logger.warning("详情规则加载失败: %s", exc)
            return {}

        sites = payload.get("sites") or {}
        rules: Dict[str, DetailRule] = {}
        for site_name, raw_rule in sites.items():
            if not isinstance(raw_rule, dict):
                continue
            rules[site_name] = DetailRule(
                content_selectors=_ensure_string_list(raw_rule.get("content_selectors")),
                remove_selectors=_ensure_string_list(raw_rule.get("remove_selectors")),
                attachment_selectors=_ensure_string_list(raw_rule.get("attachment_selectors")),
                prefer_readability=bool(raw_rule.get("prefer_readability", False)),
                force_browser_fetch=(str(raw_rule.get("force_browser_fetch")).strip() or None)
                if raw_rule.get("force_browser_fetch") is not None
                else None,
            )
        return rules

    def get_rule(self, site_name: Optional[str]) -> DetailRule:
        if not site_name:
            return DetailRule()
        return self.rules.get(site_name, DetailRule())

    async def process_jobs(
        self,
        jobs_data: List[Dict[str, Any]],
        run_time: Optional[datetime] = None,
    ) -> Dict[str, Any]:
        if not jobs_data:
            return {
                "generated": False,
                "reason": "no_jobs",
                "processed_count": 0,
                "output_root": str(self.output_root),
                "review_output_root": str(self.review_output_root),
                "results": [],
            }

        effective_run_time = run_time or datetime.now()
        run_id = effective_run_time.strftime("%Y%m%d_%H%M%S")
        run_dir = self.output_root / run_id
        run_dir.mkdir(parents=True, exist_ok=True)

        processed_jobs = list(jobs_data[: self.max_items]) if self.max_items else list(jobs_data)
        results = []
        detail_success_count = 0
        wechat_ready_count = 0

        for job in processed_jobs:
            result = await self.process_job(job, run_dir, effective_run_time)
            results.append(result)
            if result.get("detail_success"):
                detail_success_count += 1
            if result.get("wechat_ready"):
                wechat_ready_count += 1

        summary = {
            "generated": detail_success_count > 0,
            "processed_count": len(processed_jobs),
            "detail_success_count": detail_success_count,
            "wechat_ready_count": wechat_ready_count,
            "success_count": wechat_ready_count,
            "failed_count": len(processed_jobs) - detail_success_count,
            "wechat_not_ready_count": max(detail_success_count - wechat_ready_count, 0),
            "output_root": str(run_dir),
            "review_output_root": str(self.review_output_root / run_id),
            "results": results,
        }

        summary_path = run_dir / "summary.json"
        summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        summary["summary_path"] = str(summary_path)
        return summary

    async def process_job(
        self,
        job: Dict[str, Any],
        run_dir: Path,
        run_time: datetime,
    ) -> Dict[str, Any]:
        title = (job.get("title") or "未命名公告").strip()
        url = (job.get("url") or "").strip()
        site_name = (job.get("source_site") or "").strip()
        hospital = (job.get("hospital") or "未知医院").strip()

        slug = _build_slug(title, url)
        article_dir = _dedupe_directory(run_dir / slug)
        assets_dir = article_dir / "assets"
        inline_dir = assets_dir / "inline"
        attachments_dir = assets_dir / "attachments"
        renders_dir = assets_dir / "renders"
        inline_dir.mkdir(parents=True, exist_ok=True)
        attachments_dir.mkdir(parents=True, exist_ok=True)
        renders_dir.mkdir(parents=True, exist_ok=True)

        manifest: Dict[str, Any] = {
            "title": title,
            "url": url,
            "hospital": hospital,
            "source_site": site_name,
            "processed_at": run_time.isoformat(),
            "article_dir": str(article_dir),
            "inline_images": [],
            "attachments": [],
            "rendered_images": [],
            "errors": [],
            "wechat_ready": False,
            "wechat_error": "",
        }

        rule = self.get_rule(site_name)
        runtime_config = self._build_runtime_config(job)
        raw_html, fetch_backend, fetch_failure = await self._fetch_detail_html(url, runtime_config, rule)
        manifest["fetch_backend"] = fetch_backend
        manifest["fetch_failure_category"] = fetch_failure

        if not raw_html:
            failure_reason = fetch_failure or "failed_to_fetch_detail"
            manifest["errors"].append(failure_reason)
            self._write_manifest(article_dir, manifest)
            return {
                "success": False,
                "detail_success": False,
                "wechat_ready": False,
                "wechat_error": failure_reason,
                "status": "detail_fetch_failed",
                "title": title,
                "url": url,
                "article_dir": str(article_dir),
                "manifest_path": str(article_dir / "manifest.json"),
            }

        (article_dir / "source.html").write_text(raw_html, encoding="utf-8")

        cleaned_html, cleaned_title, extraction_method = self._extract_cleaned_html(
            raw_html,
            url,
            title,
            rule,
        )
        manifest["extraction_method"] = extraction_method
        manifest["resolved_title"] = cleaned_title
        (article_dir / "cleaned.html").write_text(cleaned_html, encoding="utf-8")

        soup = BeautifulSoup(cleaned_html, "lxml")
        await self._localize_inline_images(soup, inline_dir, runtime_config, manifest, article_dir)
        attachments = self._discover_attachments(soup, url, rule, raw_html=raw_html)
        await self._download_and_render_attachments(
            attachments,
            attachments_dir,
            renders_dir,
            runtime_config,
            manifest,
            article_dir,
        )

        localized_html = str(soup)
        (article_dir / "cleaned.html").write_text(localized_html, encoding="utf-8")

        article_markdown = self._build_article_markdown(
            job=job,
            title=cleaned_title,
            localized_html=localized_html,
            manifest=manifest,
            run_time=run_time,
            article_dir=article_dir,
        )
        (article_dir / "article.md").write_text(article_markdown, encoding="utf-8")

        review_result: Dict[str, Any] = {}
        try:
            review_result = self.review_builder.build_package(
                job=job,
                title=cleaned_title,
                localized_html=localized_html,
                manifest=manifest,
                run_time=run_time,
                article_dir=article_dir,
            )
            manifest["wechat_review"] = review_result
            manifest["wechat_ready"] = True
        except Exception as exc:
            logger.warning("公众号待审包构建失败: %s", exc)
            wechat_error = f"wechat_review_build_failed:{exc}"
            manifest["wechat_error"] = wechat_error
            manifest["errors"].append(wechat_error)

        self._write_manifest(article_dir, manifest)

        wechat_ready = bool(review_result.get("package_path"))
        status = "ok" if wechat_ready else "wechat_not_ready"
        wechat_error = manifest.get("wechat_error") or ""

        return {
            "success": wechat_ready,
            "detail_success": True,
            "wechat_ready": wechat_ready,
            "wechat_error": wechat_error,
            "status": status,
            "title": cleaned_title,
            "url": url,
            "article_dir": str(article_dir),
            "article_path": str(article_dir / "article.md"),
            "manifest_path": str(article_dir / "manifest.json"),
            "review_dir": review_result.get("review_dir"),
            "review_article_html_path": review_result.get("article_html_path"),
            "review_package_path": review_result.get("package_path"),
            "article_type": review_result.get("article_type"),
        }

    def _build_runtime_config(self, job: Dict[str, Any]) -> CrawlerConfig:
        site_name = (job.get("source_site") or "").strip()
        url = (job.get("url") or "").strip()
        existing = self.crawler.site_configs.get(site_name) if hasattr(self.crawler, "site_configs") else None
        if existing:
            return CrawlerConfig(
                {
                    "site_name": existing.site_name,
                    "base_url": existing.base_url,
                    "start_url": url,
                    "enabled": existing.enabled,
                    "priority": existing.priority,
                    "request_timeout": existing.request_timeout,
                    "max_retries": existing.max_retries,
                    "delay_between_requests": existing.delay_between_requests,
                    "user_agent": existing.user_agent,
                    "parsing": existing.parsing,
                    "waf_strategy": existing.waf_strategy,
                    "validation": existing.validation,
                    "ssl_verify": existing.ssl_verify,
                }
            )

        parsed = urlparse(url)
        base_url = f"{parsed.scheme}://{parsed.netloc}" if parsed.scheme and parsed.netloc else url
        return CrawlerConfig(
            {
                "site_name": site_name or "detail_pipeline",
                "base_url": base_url,
                "start_url": url,
                "request_timeout": 30,
                "max_retries": 3,
                "delay_between_requests": 0,
                "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "parsing": {},
                "waf_strategy": {},
                "validation": {},
                "ssl_verify": True,
            }
        )

    async def _fetch_detail_html(
        self,
        url: str,
        config: CrawlerConfig,
        rule: DetailRule,
    ) -> Tuple[Optional[str], str, str]:
        fetch_backend = "crawler"
        content: Optional[str] = None
        failure_category = ""

        if (rule.force_browser_fetch or "").lower() == "scrapling":
            content = await self._fetch_with_scrapling(url)
            fetch_backend = "scrapling"

        if not content:
            fetch_result = await self.crawler.fetch_page_with_metadata(url, config)
            content = fetch_result.get("content")
            fetch_backend = str(fetch_result.get("fetch_backend") or "crawler")
            failure_category = str(fetch_result.get("failure_category") or "")

        if content and self._looks_like_blocked_page(content):
            fallback_content = await self._fetch_with_scrapling(url)
            if fallback_content:
                content = fallback_content
                fetch_backend = "scrapling"
                failure_category = ""

        return content, fetch_backend, failure_category

    @staticmethod
    def _looks_like_blocked_page(content: str) -> bool:
        if not content:
            return False
        normalized = content.lower()
        if _looks_like_upstream_error_page(content):
            return True
        if _looks_like_yunsuo_verify_page(content) or _looks_like_safeline_js_challenge(content):
            return True
        markers = (
            "captcha",
            "robot check",
            "access denied",
            "verify you are human",
            "人机验证",
            "访问受限",
        )
        return any(marker in normalized for marker in markers)

    async def _fetch_with_scrapling(self, url: str) -> Optional[str]:
        if StealthyFetcher is None and DynamicFetcher is None:
            return None

        def _run_fetch() -> Optional[str]:
            fetcher_cls = StealthyFetcher or DynamicFetcher
            if fetcher_cls is None:
                return None
            try:
                fetcher = fetcher_cls()
                response = fetcher.get(url)
                return getattr(response, "text", None) or str(response)
            except Exception as exc:
                logger.warning("Scrapling 抓取失败: %s", exc)
                return None

        return await asyncio.to_thread(_run_fetch)

    def _extract_cleaned_html(
        self,
        raw_html: str,
        url: str,
        fallback_title: str,
        rule: DetailRule,
    ) -> Tuple[str, str, str]:
        extraction_order = []
        if rule.prefer_readability:
            extraction_order.extend(("readability", "selectors", "body"))
        else:
            extraction_order.extend(("selectors", "readability", "body"))

        chosen_html = ""
        extraction_method = "body"

        for method in extraction_order:
            if method == "selectors":
                chosen_html = self._extract_with_selectors(raw_html, rule.content_selectors)
            elif method == "readability":
                chosen_html = self._extract_with_readability(raw_html)
            else:
                chosen_html = self._extract_from_body(raw_html)

            if _has_meaningful_text(chosen_html):
                extraction_method = method
                break

        soup = BeautifulSoup(chosen_html or self._extract_from_body(raw_html), "lxml")
        self._prune_html(soup, rule)
        self._absolutize_links(soup, url)
        cleaned_title = self._extract_title(raw_html, fallback_title)
        container = self._ensure_article_container(soup)
        cleaned_html = str(container)
        return cleaned_html, cleaned_title, extraction_method

    @staticmethod
    def _extract_with_selectors(raw_html: str, selectors: List[str]) -> str:
        if not selectors:
            return ""
        soup = BeautifulSoup(raw_html, "lxml")
        matched = []
        for selector in selectors:
            try:
                nodes = soup.select(selector)
            except Exception:
                continue
            for node in nodes:
                html = str(node)
                if _has_meaningful_text(html):
                    matched.append(html)
        return "\n".join(matched)

    @staticmethod
    def _extract_with_readability(raw_html: str) -> str:
        if Document is None:
            return ""
        try:
            doc = Document(raw_html)
            return doc.summary(html_partial=True)
        except Exception as exc:
            logger.warning("readability 提取失败: %s", exc)
            return ""

    @staticmethod
    def _extract_from_body(raw_html: str) -> str:
        soup = BeautifulSoup(raw_html, "lxml")
        body = soup.body or soup
        return str(body)

    @staticmethod
    def _extract_title(raw_html: str, fallback_title: str) -> str:
        soup = BeautifulSoup(raw_html, "lxml")
        cleaned_fallback_title = _clean_title_text(fallback_title)
        for selector in (
            ".content-title",
            ".announcement_show_tit",
            ".article-title",
            ".news-title",
            ".title",
            "h1",
        ):
            node = soup.select_one(selector)
            if node and node.get_text(strip=True):
                candidate_title = _clean_title_text(node.get_text(strip=True))
                return _select_better_detail_title(
                    extracted_title=candidate_title,
                    fallback_title=cleaned_fallback_title,
                )
        if soup.title and soup.title.get_text(strip=True):
            candidate_title = _clean_title_text(soup.title.get_text(strip=True))
            return _select_better_detail_title(
                extracted_title=candidate_title,
                fallback_title=cleaned_fallback_title,
            )
        return cleaned_fallback_title

    def _prune_html(self, soup: BeautifulSoup, rule: DetailRule) -> None:
        for selector in DEFAULT_REMOVE_SELECTORS + rule.remove_selectors:
            try:
                for node in soup.select(selector):
                    node.decompose()
            except Exception:
                continue

        for tag_name in ("script", "style", "nav", "footer", "aside", "iframe", "form", "noscript"):
            for node in soup.find_all(tag_name):
                node.decompose()

        for node in soup.find_all(True):
            attrs = getattr(node, "attrs", None)
            classes = " ".join(attrs.get("class", [])) if isinstance(attrs, dict) else ""
            if re.search(r"(share|sidebar|advert|ads|breadcrumb|footer|header)", classes, re.I):
                node.decompose()

        for node in list(soup.find_all(["p", "div", "span"])):
            if _looks_like_hidden_text_node(node):
                node.decompose()
                continue
            text = node.get_text(" ", strip=True)
            if not text and not node.find("img") and not node.find("table"):
                node.decompose()
                continue
            if any(re.search(pattern, text) for pattern in TRAILING_NOISE_PATTERNS):
                node.decompose()

    @staticmethod
    def _absolutize_links(soup: BeautifulSoup, base_url: str) -> None:
        for tag in soup.find_all(["a", "img", "source"]):
            for attr in ("href", "src", "data-src"):
                if not tag.get(attr):
                    continue
                value = tag.get(attr).strip()
                if value.startswith("data:"):
                    continue
                tag[attr] = urljoin(base_url, value)

    @staticmethod
    def _ensure_article_container(soup: BeautifulSoup) -> Tag:
        if soup.body:
            children = [child for child in soup.body.children if getattr(child, "name", None) or str(child).strip()]
        else:
            children = [child for child in soup.children if getattr(child, "name", None) or str(child).strip()]

        wrapper = BeautifulSoup("<article></article>", "lxml").article
        assert wrapper is not None
        for child in children:
            wrapper.append(child)
        return wrapper

    async def _localize_inline_images(
        self,
        soup: BeautifulSoup,
        inline_dir: Path,
        config: CrawlerConfig,
        manifest: Dict[str, Any],
        article_dir: Path,
    ) -> None:
        image_index = 0
        for img in soup.find_all("img"):
            src = (img.get("src") or img.get("data-src") or "").strip()
            if not src:
                continue
            image_index += 1
            filename = _build_resource_name(src, image_index, "inline")
            destination = inline_dir / filename
            downloaded, content_type = await self._download_binary_resource(src, destination, config)
            if not downloaded:
                manifest["errors"].append(f"inline_image_download_failed:{src}")
                continue
            relative_path = destination.relative_to(article_dir).as_posix()
            img["src"] = relative_path
            manifest["inline_images"].append(
                {
                    "source_url": src,
                    "local_path": relative_path,
                    "content_type": content_type,
                }
            )

    def _discover_attachments(
        self,
        soup: BeautifulSoup,
        base_url: str,
        rule: DetailRule,
        raw_html: str = "",
    ) -> List[Dict[str, Any]]:
        attachments: List[Dict[str, Any]] = []
        seen_urls = set()

        explicit_nodes = self._collect_attachment_nodes(soup, rule.attachment_selectors)
        nodes: List[Tag] = explicit_nodes or list(soup.find_all("a", href=True))

        if raw_html:
            raw_soup = BeautifulSoup(raw_html, "lxml")
            raw_explicit_nodes = self._collect_attachment_nodes(raw_soup, rule.attachment_selectors)
            raw_fallback_nodes = self._collect_attachment_nodes(raw_soup, RAW_ATTACHMENT_SELECTORS)
            for node in raw_explicit_nodes + raw_fallback_nodes:
                if node not in nodes:
                    nodes.append(node)

        for index, node in enumerate(nodes, start=1):
            href = urljoin(base_url, (node.get("href") or "").strip())
            if not href or href in seen_urls:
                continue
            text = node.get_text(" ", strip=True)
            ext = _guess_extension(href)
            is_attachment = node in explicit_nodes or ext in ATTACHMENT_EXTENSIONS or any(
                keyword in text for keyword in ATTACHMENT_HINT_KEYWORDS
            )
            if not is_attachment:
                continue
            seen_urls.add(href)
            attachments.append(
                {
                    "index": index,
                    "source_url": href,
                    "title": text or Path(urlparse(href).path).name or f"attachment_{index}",
                    "extension": ext,
                }
            )
        return attachments

    @staticmethod
    def _collect_attachment_nodes(soup: BeautifulSoup, selectors: List[str] | tuple[str, ...]) -> List[Tag]:
        nodes: List[Tag] = []
        seen: set[int] = set()
        for selector in selectors:
            try:
                matched = soup.select(selector)
            except Exception:
                continue
            for node in matched:
                identifier = id(node)
                if identifier in seen:
                    continue
                seen.add(identifier)
                nodes.append(node)
        return nodes

    async def _download_and_render_attachments(
        self,
        attachments: List[Dict[str, Any]],
        attachments_dir: Path,
        renders_dir: Path,
        config: CrawlerConfig,
        manifest: Dict[str, Any],
        article_dir: Path,
    ) -> None:
        for attachment in attachments:
            filename = _build_resource_name(
                attachment["source_url"],
                attachment["index"],
                "attachment",
                preferred_extension=attachment.get("extension") or None,
            )
            destination = attachments_dir / filename
            downloaded, content_type = await self._download_binary_resource(
                attachment["source_url"],
                destination,
                config,
            )
            attachment_result = {
                "title": attachment["title"],
                "source_url": attachment["source_url"],
                "content_type": content_type,
                "local_path": destination.relative_to(article_dir).as_posix() if downloaded else None,
                "category": _classify_attachment_type(attachment["title"], attachment["source_url"]),
                "render_status": "skipped",
                "rendered_images": [],
            }

            if not downloaded:
                attachment_result["render_status"] = "download_failed"
                manifest["attachments"].append(attachment_result)
                manifest["errors"].append(f"attachment_download_failed:{attachment['source_url']}")
                continue

            render_paths, render_status = await self._render_attachment(
                destination,
                attachment_result,
                renders_dir,
            )
            attachment_result["render_status"] = render_status
            attachment_result["rendered_images"] = [
                render_path.relative_to(article_dir).as_posix() for render_path in render_paths
            ]
            manifest["attachments"].append(attachment_result)
            manifest["rendered_images"].extend(attachment_result["rendered_images"])

    async def _render_attachment(
        self,
        attachment_path: Path,
        attachment: Dict[str, Any],
        renders_dir: Path,
    ) -> Tuple[List[Path], str]:
        if attachment.get("category") != "position_table":
            return [], "link_only"
        ext = attachment_path.suffix.lower()
        if ext not in RENDERABLE_ATTACHMENT_EXTENSIONS:
            return [], "unsupported"

        if ext == ".pdf":
            return await asyncio.to_thread(self._render_pdf, attachment_path, renders_dir)
        if ext == ".xlsx":
            return await asyncio.to_thread(self._render_xlsx, attachment_path, renders_dir)
        if ext == ".doc":
            return await asyncio.to_thread(self._render_doc, attachment_path, renders_dir)
        if ext == ".docx":
            return await asyncio.to_thread(self._render_docx, attachment_path, renders_dir)
        return [], "unsupported"

    def _render_doc(self, attachment_path: Path, renders_dir: Path) -> Tuple[List[Path], str]:
        table_render = self._render_doc_table_with_textutil(attachment_path, renders_dir)
        if table_render[0]:
            return table_render

        converted_path = self._convert_legacy_doc_to_docx(attachment_path)
        if converted_path is None:
            return [], "unsupported"
        try:
            return self._render_docx(converted_path, renders_dir)
        finally:
            try:
                converted_path.unlink(missing_ok=True)
            except Exception:
                pass

    def _render_doc_table_with_textutil(
        self,
        attachment_path: Path,
        renders_dir: Path,
    ) -> Tuple[List[Path], str]:
        textutil = shutil.which("textutil")
        if not textutil:
            return [], "unsupported"

        temp_dir = Path(tempfile.mkdtemp(prefix="legacy_doc_text_"))
        txt_path = temp_dir / f"{attachment_path.stem}.txt"
        try:
            subprocess.run(
                [textutil, "-convert", "txt", "-output", str(txt_path), str(attachment_path)],
                check=True,
                capture_output=True,
                text=True,
            )
            if not txt_path.exists():
                return [], "unsupported"

            doc_title, rows = _extract_legacy_doc_table_rows_from_text(
                txt_path.read_text(encoding="utf-8", errors="ignore")
            )
            if len(rows) <= 1:
                return [], "unsupported"

            rendered_paths: List[Path] = []
            header = rows[0]
            data_rows = rows[1:]
            chunk_size = 6 if len(data_rows) > 6 else len(data_rows)
            for chunk_index, chunk in enumerate(_chunk_rows(data_rows, chunk_size), start=1):
                output_path = renders_dir / f"{attachment_path.stem}_{chunk_index}.png"
                self._render_table_rows_to_image(
                    title=doc_title or attachment_path.stem,
                    rows=[header, *chunk],
                    output_path=output_path,
                )
                rendered_paths.append(output_path)
            return rendered_paths, "rendered" if rendered_paths else "empty"
        except Exception as exc:
            logger.warning("textutil DOC 表格转图失败: %s", exc)
            return [], "unsupported"
        finally:
            try:
                txt_path.unlink(missing_ok=True)
            except Exception:
                pass
            try:
                temp_dir.rmdir()
            except Exception:
                pass

    def _convert_legacy_doc_to_docx(self, attachment_path: Path) -> Optional[Path]:
        temp_dir = Path(tempfile.mkdtemp(prefix="legacy_doc_render_"))
        output_path = temp_dir / f"{attachment_path.stem}.docx"

        textutil = shutil.which("textutil")
        if textutil:
            try:
                subprocess.run(
                    [textutil, "-convert", "docx", "-output", str(output_path), str(attachment_path)],
                    check=True,
                    capture_output=True,
                    text=True,
                )
                if output_path.exists():
                    return output_path
            except Exception as exc:
                logger.warning("textutil 转换 DOC 失败: %s", exc)

        for office_cmd in ("soffice", "libreoffice"):
            binary = shutil.which(office_cmd)
            if not binary:
                continue
            try:
                subprocess.run(
                    [
                        binary,
                        "--headless",
                        "--convert-to",
                        "docx",
                        "--outdir",
                        str(temp_dir),
                        str(attachment_path),
                    ],
                    check=True,
                    capture_output=True,
                    text=True,
                )
                office_output = temp_dir / f"{attachment_path.stem}.docx"
                if office_output.exists():
                    return office_output
            except Exception as exc:
                logger.warning("%s 转换 DOC 失败: %s", office_cmd, exc)

        try:
            temp_dir.rmdir()
        except Exception:
            pass
        return None

    def _render_pdf(self, attachment_path: Path, renders_dir: Path) -> Tuple[List[Path], str]:
        if fitz is None:
            return [], "missing_dependency"

        rendered_paths: List[Path] = []
        try:
            with fitz.open(attachment_path) as document:
                for page_index, page in enumerate(document, start=1):
                    pixmap = page.get_pixmap(matrix=fitz.Matrix(1.6, 1.6), alpha=False)
                    output_path = renders_dir / f"{attachment_path.stem}_page_{page_index}.png"
                    pixmap.save(output_path)
                    rendered_paths.append(output_path)
            return rendered_paths, "rendered" if rendered_paths else "empty"
        except Exception as exc:
            logger.warning("PDF 转图失败: %s", exc)
            return [], "render_failed"

    def _render_xlsx(self, attachment_path: Path, renders_dir: Path) -> Tuple[List[Path], str]:
        if load_workbook is None:
            return [], "missing_dependency"
        if Image is None or ImageDraw is None or ImageFont is None:
            return [], "missing_dependency"

        try:
            workbook = load_workbook(attachment_path, data_only=True)
        except Exception as exc:
            logger.warning("XLSX 读取失败: %s", exc)
            return [], "render_failed"

        rendered_paths: List[Path] = []
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                values = ["" if cell is None else str(cell).strip() for cell in row]
                if any(values):
                    rows.append(values)
            if not rows:
                continue

            for chunk_index, chunk in enumerate(_chunk_rows(rows, 18), start=1):
                output_path = renders_dir / f"{attachment_path.stem}_{_safe_slug(sheet.title)}_{chunk_index}.png"
                self._render_table_rows_to_image(
                    title=f"{attachment_path.stem} | {sheet.title}",
                    rows=chunk,
                    output_path=output_path,
                )
                rendered_paths.append(output_path)

        return rendered_paths, "rendered" if rendered_paths else "empty"

    def _render_docx(self, attachment_path: Path, renders_dir: Path) -> Tuple[List[Path], str]:
        if Image is None or ImageDraw is None or ImageFont is None:
            return [], "missing_dependency"

        try:
            paragraphs, tables = _extract_docx_content(attachment_path)
            rendered_paths: List[Path] = []

            if tables:
                doc_title = next((line for line in paragraphs if len(line) >= 6), attachment_path.stem)
                for table_index, rows in enumerate(tables, start=1):
                    chunk_size = 12 if len(rows) > 12 else len(rows)
                    for chunk_index, chunk in enumerate(_chunk_rows(rows, chunk_size), start=1):
                        output_path = renders_dir / f"{attachment_path.stem}_table_{table_index}_{chunk_index}.png"
                        self._render_table_rows_to_image(
                            title=doc_title,
                            rows=chunk,
                            output_path=output_path,
                        )
                        rendered_paths.append(output_path)
                if rendered_paths:
                    return rendered_paths, "rendered"

            if mammoth is not None:
                with attachment_path.open("rb") as file_obj:
                    result = mammoth.convert_to_html(file_obj)
                soup = BeautifulSoup(result.value, "lxml")
                lines = _extract_text_lines_from_html(soup)
            else:
                lines = paragraphs

            if not lines:
                return [], "empty"

            for chunk_index, chunk in enumerate(_chunk_rows([[line] for line in lines], 40), start=1):
                output_path = renders_dir / f"{attachment_path.stem}_{chunk_index}.png"
                self._render_text_lines_to_image(
                    [row[0] for row in chunk],
                    output_path,
                    title=attachment_path.stem,
                )
                rendered_paths.append(output_path)
            return rendered_paths, "rendered" if rendered_paths else "empty"
        except Exception as exc:
            logger.warning("DOCX 转图失败: %s", exc)
            return [], "render_failed"

    def _render_table_rows_to_image(
        self,
        title: str,
        rows: List[List[str]],
        output_path: Path,
    ) -> None:
        if Image is None or ImageDraw is None or ImageFont is None:
            raise RuntimeError("Pillow is required to render attachment previews")

        normalized_rows = [list(row) for row in rows if any(str(cell).strip() for cell in row)]
        if not normalized_rows:
            raise RuntimeError("没有可渲染的表格行")

        width = 1800
        margin = 40
        header_height = 86
        row_padding_y = 14
        title_font = _load_font(28)
        header_font = _load_font(22)
        body_font = _load_font(18)

        col_count = max(len(row) for row in normalized_rows)
        prepared_rows = [row + [""] * (col_count - len(row)) for row in normalized_rows]
        image = Image.new("RGB", (width, 10), "white")
        drawer = ImageDraw.Draw(image)

        col_widths = _estimate_column_widths(prepared_rows, drawer, body_font, width - margin * 2)
        row_layouts: List[List[List[str]]] = []
        row_heights: List[int] = []
        for row_index, row in enumerate(prepared_rows):
            wrapped_cells: List[List[str]] = []
            max_lines = 1
            font = header_font if row_index == 0 else body_font
            for col_index, cell in enumerate(row):
                wrapped = _wrap_text(cell or " ", font, max(col_widths[col_index] - 24, 60))
                wrapped_cells.append(wrapped)
                max_lines = max(max_lines, len(wrapped))
            row_layouts.append(wrapped_cells)
            row_heights.append(max_lines * 28 + row_padding_y * 2)

        total_height = header_height + sum(row_heights) + margin * 2
        image = Image.new("RGB", (width, total_height), "#fffdf8")
        drawer = ImageDraw.Draw(image)
        drawer.rounded_rectangle(
            (10, 10, width - 10, total_height - 10),
            radius=28,
            fill="#fffdf8",
            outline="#d9e6dd",
            width=2,
        )
        drawer.rounded_rectangle(
            (margin, margin, width - margin, margin + header_height),
            radius=20,
            fill="#eaf4ee",
            outline="#cfe0d6",
            width=2,
        )
        drawer.text((margin + 24, margin + 24), title, font=title_font, fill="#234836")

        y = margin + header_height + 20
        for row_index, wrapped_cells in enumerate(row_layouts):
            row_height = row_heights[row_index]
            x = margin
            fill = "#f4f8f5" if row_index == 0 else ("#ffffff" if row_index % 2 else "#fbfcfa")
            for col_index, cell_lines in enumerate(wrapped_cells):
                cell_width = col_widths[col_index]
                drawer.rectangle((x, y, x + cell_width, y + row_height), fill=fill, outline="#d8e4dc", width=1)
                font = header_font if row_index == 0 else body_font
                line_y = y + row_padding_y
                for line in cell_lines:
                    drawer.text((x + 12, line_y), line, font=font, fill="#25362d")
                    line_y += 28
                x += cell_width
            y += row_height
        image.save(output_path)

    def _render_text_lines_to_image(
        self,
        lines: Iterable[str],
        output_path: Path,
        title: Optional[str] = None,
    ) -> None:
        if Image is None or ImageDraw is None or ImageFont is None:
            raise RuntimeError("Pillow is required to render attachment previews")
        font = _load_font(21)
        title_font = _load_font(28)
        width = 1600
        margin = 48
        body_lines: List[Tuple[str, ImageFont.ImageFont]] = []
        title_text = (title or "").strip()
        if title_text:
            for wrapped_line in _wrap_text(title_text, title_font, width - margin * 2 - 24):
                body_lines.append((wrapped_line, title_font))
        for line in lines:
            current_font = font
            wrapped = _wrap_text(line, current_font, width - margin * 2)
            for wrapped_line in wrapped:
                body_lines.append((wrapped_line, current_font))

        line_height = 34
        header_height = 84 if title_text else 0
        height = max(320, margin * 2 + header_height + line_height * max(len(body_lines), 1))
        image = Image.new("RGB", (width, height), "#fffdf8")
        drawer = ImageDraw.Draw(image)
        drawer.rounded_rectangle(
            (10, 10, width - 10, height - 10),
            radius=28,
            fill="#fffdf8",
            outline="#d9e6dd",
            width=2,
        )
        y = margin
        if title_text:
            drawer.rounded_rectangle(
                (margin, margin, width - margin, margin + header_height),
                radius=20,
                fill="#eaf4ee",
                outline="#cfe0d6",
                width=2,
            )
            y += 22
        for line, current_font in body_lines:
            drawer.text((margin + 16, y), line, font=current_font, fill="#25362d")
            y += line_height
        image.save(output_path)

    async def _download_binary_resource(
        self,
        source_url: str,
        destination: Path,
        config: CrawlerConfig,
    ) -> Tuple[bool, Optional[str]]:
        if source_url.startswith("data:"):
            match = re.match(r"^data:([^;]+);base64,(.+)$", source_url, re.I)
            if not match:
                return False, None
            content_type, encoded = match.groups()
            destination.write_bytes(base64.b64decode(encoded))
            return True, content_type

        if not getattr(self.crawler, "session", None):
            await self.crawler.init_session()

        try:
            async with self.crawler.session.get(source_url, ssl=config.ssl_verify) as response:
                if response.status != 200:
                    return False, None
                body = await response.read()
                destination.write_bytes(body)
                return True, response.headers.get("Content-Type")
        except aiohttp.ClientError as exc:
            logger.warning("资源下载失败: %s", exc)
            return False, None

    def _build_article_markdown(
        self,
        job: Dict[str, Any],
        title: str,
        localized_html: str,
        manifest: Dict[str, Any],
        run_time: datetime,
        article_dir: Path,
    ) -> str:
        body_markdown = self._html_to_markdown(localized_html)
        inline_images = manifest.get("inline_images", [])
        attachments = manifest.get("attachments", [])

        inline_section = (
            "\n".join(
                f"{index}. ![{Path(item['local_path']).name}]({item['local_path']})"
                for index, item in enumerate(inline_images, start=1)
            )
            if inline_images
            else "无"
        )

        attachment_notes = []
        render_lines = []
        original_lines = []
        for index, attachment in enumerate(attachments, start=1):
            local_path = attachment.get("local_path")
            title_text = attachment.get("title") or Path(local_path or "").name or f"附件{index}"
            if local_path:
                original_lines.append(f"{index}. [{title_text}]({local_path})")
            else:
                original_lines.append(f"{index}. {title_text}")

            render_status = attachment.get("render_status")
            if render_status == "rendered":
                attachment_notes.append(f"- {title_text}：已下载并转成图片")
                render_lines.append(f"### {title_text}")
                for image_path in attachment.get("rendered_images", []):
                    render_lines.append(f"![{title_text}]({image_path})")
            elif render_status == "link_only":
                attachment_notes.append(f"- {title_text}：按规则保留下载链接，不转图")
            elif render_status == "unsupported":
                attachment_notes.append(f"- {title_text}：附件已下载，当前格式未自动转图")
            elif render_status == "missing_dependency":
                attachment_notes.append(f"- {title_text}：附件已下载，但缺少转换依赖，未转图")
            elif render_status == "download_failed":
                attachment_notes.append(f"- {title_text}：附件下载失败")
            else:
                attachment_notes.append(f"- {title_text}：附件已下载，未转图")

        attachment_notes_text = "\n".join(attachment_notes) if attachment_notes else "无"
        rendered_section = "\n\n".join(render_lines) if render_lines else "无"
        original_section = "\n".join(original_lines) if original_lines else "无"

        relative_manifest = (article_dir / "manifest.json").name
        return (
            f"# {title}\n\n"
            f"- 医院：{job.get('hospital') or '未知医院'}\n"
            f"- 来源链接：{job.get('url')}\n"
            f"- 抓取时间：{run_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"- 来源站点：{job.get('source_site') or 'unknown'}\n"
            f"- 清单文件：`{relative_manifest}`\n\n"
            f"## 正文\n\n"
            f"{body_markdown.strip() or '无正文内容'}\n\n"
            f"## 正文图片\n\n"
            f"{inline_section}\n\n"
            f"## 附件说明\n\n"
            f"{attachment_notes_text}\n\n"
            f"## 附件转图片内容\n\n"
            f"{rendered_section}\n\n"
            f"## 原附件清单\n\n"
            f"{original_section}\n"
        )

    def _html_to_markdown(self, html: str) -> str:
        if html_to_markdown is not None:
            try:
                return html_to_markdown(
                    html,
                    heading_style="ATX",
                    bullets="-",
                    strip=["span"],
                )
            except Exception as exc:
                logger.warning("markdownify 转换失败，改用简易转换: %s", exc)

        soup = BeautifulSoup(html, "lxml")
        container = soup.body or soup
        blocks = []
        for child in container.children:
            rendered = self._render_markdown_block(child)
            if rendered:
                blocks.append(rendered.strip())
        return "\n\n".join(blocks).strip()

    def _render_markdown_block(self, node: Any) -> str:
        if isinstance(node, NavigableString):
            text = str(node).strip()
            return text
        if not isinstance(node, Tag):
            return ""

        name = node.name.lower()
        if name in {"h1", "h2", "h3", "h4"}:
            level = int(name[1])
            return f"{'#' * level} {self._render_inline_markdown(node).strip()}"
        if name in {"p", "div", "section", "article"}:
            if node.find("table"):
                return "\n\n".join(self._render_markdown_block(child) for child in node.children)
            return self._render_inline_markdown(node).strip()
        if name == "img":
            src = node.get("src", "").strip()
            alt = node.get("alt", "").strip() or Path(src).name
            return f"![{alt}]({src})" if src else ""
        if name == "ul":
            return "\n".join(
                f"- {self._render_inline_markdown(li).strip()}" for li in node.find_all("li", recursive=False)
            )
        if name == "ol":
            return "\n".join(
                f"{index}. {self._render_inline_markdown(li).strip()}"
                for index, li in enumerate(node.find_all("li", recursive=False), start=1)
            )
        if name == "table":
            return _table_to_markdown(node)
        if name == "br":
            return "\n"
        return "\n\n".join(
            rendered for rendered in (self._render_markdown_block(child) for child in node.children) if rendered
        )

    def _render_inline_markdown(self, node: Tag) -> str:
        fragments = []
        for child in node.children:
            if isinstance(child, NavigableString):
                text = str(child)
                if text:
                    fragments.append(text)
                continue
            if not isinstance(child, Tag):
                continue

            name = child.name.lower()
            if name == "a":
                href = (child.get("href") or "").strip()
                text = child.get_text(" ", strip=True) or href
                fragments.append(f"[{text}]({href})" if href else text)
            elif name == "img":
                src = (child.get("src") or "").strip()
                alt = child.get("alt", "").strip() or Path(src).name
                if src:
                    fragments.append(f"![{alt}]({src})")
            elif name == "br":
                fragments.append("\n")
            else:
                fragments.append(self._render_inline_markdown(child))

        combined = "".join(fragments)
        combined = re.sub(r"[ \t]+\n", "\n", combined)
        combined = re.sub(r"\n{3,}", "\n\n", combined)
        return combined.strip()

    @staticmethod
    def _write_manifest(article_dir: Path, manifest: Dict[str, Any]) -> None:
        (article_dir / "manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )


def _ensure_string_list(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, str):
        return [value.strip()] if value.strip() else []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return []


def _has_meaningful_text(html: str) -> bool:
    if not html:
        return False
    text = BeautifulSoup(html, "lxml").get_text(" ", strip=True)
    return len(re.sub(r"\s+", "", text)) >= 30


def _safe_slug(value: str) -> str:
    normalized = re.sub(r"[^a-zA-Z0-9]+", "_", value.strip().lower()).strip("_")
    return normalized[:80] or "item"


def _build_slug(title: str, url: str) -> str:
    digest = hashlib.sha1(url.encode("utf-8")).hexdigest()[:10]
    title_slug = _safe_slug(title)
    return f"{title_slug}_{digest}"


def _dedupe_directory(path: Path) -> Path:
    if not path.exists():
        path.mkdir(parents=True, exist_ok=True)
        return path
    for suffix in range(2, 1000):
        candidate = path.with_name(f"{path.name}_{suffix}")
        if not candidate.exists():
            candidate.mkdir(parents=True, exist_ok=True)
            return candidate
    raise RuntimeError(f"无法创建唯一目录: {path}")


def _guess_extension(url: str) -> str:
    parsed = urlparse(url)
    path = unquote(parsed.path or "")
    suffix = Path(path).suffix.lower()
    if suffix:
        return suffix
    query = unquote(parsed.query or "").lower()
    for ext in ATTACHMENT_EXTENSIONS | INLINE_IMAGE_EXTENSIONS:
        if ext.lstrip(".") in query:
            return ext
    return ""


def _clean_title_text(value: str) -> str:
    title = re.sub(r"\s+", " ", (value or "").strip())
    if not title:
        return ""
    title = re.sub(r"^[>＞]+\s*", "", title).strip()
    title = re.sub(r"返回列表\s*[>＞≫]*", "", title).strip()
    title = re.split(r"(?:发布时间|发布日期|发布人|访问人数|浏览次数)[:：]", title, maxsplit=1)[0].strip()
    title = re.sub(r"\s*\d{4}[./-]\d{1,2}[./-]\d{1,2}$", "", title).strip()
    suffix_keywords = ("官方网站", "医院公告", "招贤纳士", "新闻中心", "人事招聘", "招聘公告", "通知公告")
    if "-" in title:
        parts = [part.strip() for part in title.split("-") if part.strip()]
        if len(parts) >= 2 and (
            len(parts) >= 3 or any(any(keyword in part for keyword in suffix_keywords) for part in parts[1:])
        ):
            title = parts[0]
        elif len(parts) >= 2 and all(len(part) <= 24 for part in parts[1:]):
            title = parts[0]
    separators = (" - ", "——", "__", "|")
    for separator in separators:
        if separator in title:
            parts = [part.strip() for part in title.split(separator) if part.strip()]
            if parts:
                title = parts[0]
                break
    return title.strip()


def _title_has_recruitment_signal(title: str) -> bool:
    return any(keyword in (title or "") for keyword in ("招聘", "招募", "公告", "简章", "公示", "启示", "报名"))


def _looks_generic_detail_title(title: str) -> bool:
    normalized = (title or "").strip()
    if not normalized:
        return True
    compact = re.sub(r"\s+", "", normalized)
    generic_literals = {
        "人才招聘",
        "人才招聘talentrecruitment",
        "talentrecruitment",
        "招聘信息",
        "人事招聘",
        "通知公告",
        "公告信息",
        "招标招聘",
    }
    if compact.lower() in generic_literals:
        return True
    if compact.endswith("医院") and not _title_has_recruitment_signal(compact):
        return True
    if len(compact) <= 12 and not _title_has_recruitment_signal(compact):
        return True
    return False


def _select_better_detail_title(extracted_title: str, fallback_title: str) -> str:
    extracted = _clean_title_text(extracted_title)
    fallback = _clean_title_text(fallback_title)
    if not fallback:
        return extracted
    if not extracted:
        return fallback
    if _looks_generic_detail_title(extracted):
        return fallback
    if _title_has_recruitment_signal(fallback) and not _title_has_recruitment_signal(extracted):
        return fallback
    if len(fallback) >= len(extracted) + 6 and _title_has_recruitment_signal(fallback):
        return fallback
    return extracted


def _build_resource_name(
    source_url: str,
    index: int,
    prefix: str,
    preferred_extension: Optional[str] = None,
) -> str:
    parsed = urlparse(source_url)
    original_name = Path(unquote(parsed.path or "")).name
    extension = preferred_extension or Path(original_name).suffix.lower()
    if not extension:
        guessed_extension = mimetypes.guess_extension(mimetypes.guess_type(source_url)[0] or "")
        extension = guessed_extension or ".bin"
    digest = hashlib.sha1(source_url.encode("utf-8")).hexdigest()[:10]
    base_name = _safe_slug(Path(original_name).stem or f"{prefix}_{index}")
    return f"{prefix}_{index:02d}_{base_name}_{digest}{extension}"


def _chunk_rows(rows: List[List[str]], size: int) -> Iterable[List[List[str]]]:
    for start in range(0, len(rows), size):
        yield rows[start : start + size]


def _extract_text_lines_from_html(soup: BeautifulSoup) -> List[str]:
    lines = []
    for node in soup.find_all(["h1", "h2", "h3", "h4", "p", "li"]):
        text = node.get_text(" ", strip=True)
        if text:
            lines.append(text)
    if lines:
        return lines
    fallback_text = soup.get_text("\n", strip=True)
    return [line.strip() for line in fallback_text.splitlines() if line.strip()]


def _extract_legacy_doc_table_rows_from_text(text: str) -> Tuple[str, List[List[str]]]:
    normalized = (text or "").replace("\r\n", "\n").replace("\r", "\n").replace("\u2028", "\n")
    lines = [line.strip() for line in normalized.splitlines() if line.strip()]
    if not lines:
        return "", []

    table_start = next((index for index, line in enumerate(lines) if "\x07" in line), -1)
    if table_start < 0:
        return "", []

    title_lines = [line for line in lines[:table_start] if line]
    doc_title = next((line for line in reversed(title_lines) if len(line) >= 6), title_lines[-1] if title_lines else "")
    table_text = "\n".join(lines[table_start:])
    chunks = [chunk for chunk in re.split(r"\x07{2,}(?=\d+\x07)", table_text) if chunk.strip()]
    if len(chunks) < 2:
        return doc_title, []

    data_rows = [_split_legacy_doc_row(chunk) for chunk in chunks[1:]]
    column_count = _infer_legacy_doc_column_count(data_rows)
    if column_count < 4:
        return doc_title, []

    header_row = _build_legacy_doc_header_row(chunks[0], column_count)
    rows = [header_row]
    merged_unit_value = ""
    for chunk in chunks[1:]:
        cells = _split_legacy_doc_row(chunk)
        if _looks_like_legacy_doc_page_marker(cells):
            continue
        cells = _normalize_legacy_doc_row(cells, column_count)
        if not cells:
            continue
        if column_count >= 2:
            if cells[1]:
                merged_unit_value = cells[1]
            elif merged_unit_value:
                cells[1] = merged_unit_value
        rows.append(cells)

    return doc_title, rows


def _split_legacy_doc_row(chunk: str) -> List[str]:
    cells = [part.strip() for part in chunk.split("\x07")]
    while cells and not cells[0]:
        cells.pop(0)
    while cells and not cells[-1]:
        cells.pop()
    return [re.sub(r"\n{3,}", "\n\n", cell) for cell in cells]


def _build_legacy_doc_header_row(header_chunk: str, column_count: int) -> List[str]:
    if column_count == len(LEGACY_DOC_STANDARD_HEADERS):
        return LEGACY_DOC_STANDARD_HEADERS

    tokens = [token for token in _split_legacy_doc_row(header_chunk) if token]
    if (
        column_count == len(LEGACY_DOC_STANDARD_HEADERS)
        and "岗位条件" in tokens
        and "执业（职称）条件" in tokens
        and "其他条件" in tokens
    ):
        return LEGACY_DOC_STANDARD_HEADERS
    if len(tokens) >= column_count:
        return tokens[:column_count]
    return tokens + [""] * max(column_count - len(tokens), 0)


def _infer_legacy_doc_column_count(rows: List[List[str]]) -> int:
    counts = Counter(
        len(row)
        for row in rows
        if row and not _looks_like_legacy_doc_page_marker(row) and len(row) >= 5
    )
    if counts:
        return counts.most_common(1)[0][0]
    return max((len(row) for row in rows), default=0)


def _looks_like_legacy_doc_page_marker(cells: List[str]) -> bool:
    return any(LEGACY_DOC_PAGE_MARKER_PATTERN.search(cell or "") for cell in cells)


def _normalize_legacy_doc_row(cells: List[str], column_count: int) -> List[str]:
    cleaned = [_strip_legacy_doc_noise(cell) for cell in cells]
    while cleaned and not cleaned[-1]:
        cleaned.pop()
    if not cleaned or _looks_like_legacy_doc_page_marker(cleaned):
        return []
    if len(cleaned) > column_count:
        head = cleaned[: column_count - 1]
        tail = [cell for cell in cleaned[column_count - 1 :] if cell]
        cleaned = head + ["\n".join(tail)]
    elif len(cleaned) < column_count:
        cleaned.extend([""] * (column_count - len(cleaned)))
    return cleaned


def _strip_legacy_doc_noise(value: str) -> str:
    normalized = (value or "").strip()
    normalized = LEGACY_DOC_PAGE_MARKER_PATTERN.sub("", normalized)
    normalized = re.sub(r"\n{3,}", "\n\n", normalized)
    return normalized.strip()


def _extract_docx_content(docx_path: Path) -> Tuple[List[str], List[List[List[str]]]]:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    try:
        with zipfile.ZipFile(docx_path) as archive:
            xml = archive.read("word/document.xml")
    except Exception as exc:
        raise RuntimeError(f"DOCX 读取失败: {exc}") from exc

    root = ET.fromstring(xml)
    paragraphs: List[str] = []
    for paragraph in root.findall(".//w:p", ns):
        text = "".join(node.text or "" for node in paragraph.findall(".//w:t", ns)).strip()
        if text:
            paragraphs.append(re.sub(r"\s+", " ", text))

    tables: List[List[List[str]]] = []
    for table in root.findall(".//w:tbl", ns):
        rows: List[List[str]] = []
        for tr in table.findall("./w:tr", ns):
            row: List[str] = []
            for tc in tr.findall("./w:tc", ns):
                cell_text = "".join(node.text or "" for node in tc.findall(".//w:t", ns)).strip()
                row.append(re.sub(r"\s+", " ", cell_text))
            if any(cell.strip() for cell in row):
                while row and not row[-1].strip():
                    row.pop()
                rows.append(row)
        if rows:
            tables.append(rows)

    return paragraphs, tables


def _looks_like_hidden_text_node(node: Tag) -> bool:
    attrs = getattr(node, "attrs", None)
    if not isinstance(attrs, dict):
        return False
    style = (attrs.get("style") or "").lower()
    text = re.sub(r"\s+", "", node.get_text(" ", strip=True))
    if not text or node.find(["img", "table"]):
        return False
    if "display:none" in style or "visibility:hidden" in style or "opacity:0" in style or "mso-hide:all" in style:
        return True
    if not _has_white_text_style(style):
        return False
    font_size = _extract_inline_font_size(style)
    if font_size is not None and font_size <= 14:
        return True
    return len(text) <= 8


def _has_white_text_style(style: str) -> bool:
    return any(
        marker in style
        for marker in (
            "color:#fff",
            "color: #fff",
            "color:#ffffff",
            "color: #ffffff",
            "color:white",
            "color: white",
            "color:rgb(255,255,255)",
            "color: rgb(255,255,255)",
        )
    )


def _extract_inline_font_size(style: str) -> Optional[float]:
    match = re.search(r"font-size\s*:\s*([0-9.]+)\s*(px|pt)?", style)
    if not match:
        return None
    size = float(match.group(1))
    unit = (match.group(2) or "px").lower()
    if unit == "pt":
        return size
    return size * 0.75


def _table_to_markdown(table: Tag) -> str:
    rows = []
    for tr in table.find_all("tr"):
        cells = [cell.get_text(" ", strip=True) for cell in tr.find_all(["th", "td"])]
        if any(cells):
            rows.append(cells)
    if not rows:
        return ""

    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    header = normalized[0]
    separator = ["---"] * width
    body = normalized[1:] or []

    markdown_rows = [
        f"| {' | '.join(header)} |",
        f"| {' | '.join(separator)} |",
    ]
    markdown_rows.extend(f"| {' | '.join(row)} |" for row in body)
    return "\n".join(markdown_rows)


def _classify_attachment_type(title: str, source_url: str = "") -> str:
    normalized = f"{title} {source_url}".lower()
    if any(keyword in normalized for keyword in POSITION_ATTACHMENT_KEYWORDS):
        return "position_table"
    if any(keyword in normalized for keyword in FORM_ATTACHMENT_KEYWORDS):
        return "application_form"
    return "other"


def _estimate_column_widths(
    rows: List[List[str]],
    drawer: Any,
    font: Any,
    total_width: int,
) -> List[int]:
    if not rows:
        return [total_width]

    col_count = max(len(row) for row in rows)
    headers = [rows[0][index].strip() if rows and index < len(rows[0]) else "" for index in range(col_count)]
    preset_widths = _preset_column_widths(headers, total_width)
    if preset_widths is not None:
        return preset_widths
    weights = [_column_weight_hint(header) for header in headers]
    for row in rows[: min(len(rows), 12)]:
        for index in range(col_count):
            cell = row[index] if index < len(row) else ""
            header = headers[index]
            text = (cell or " ").strip()
            category = _classify_column_text_density(header)
            sample_limit = 56 if category == "long" else 28
            text = (text[:sample_limit] or " ").strip() or " "
            try:
                bbox = drawer.textbbox((0, 0), text, font=font)
                pixel_width = bbox[2] - bbox[0]
            except Exception:
                pixel_width = len(text) * 18
            max_weight = 20.0 if category == "long" else (8.0 if category == "short" else 12.0)
            weights[index] = max(weights[index], min(pixel_width / 18.0, max_weight))

    min_widths = [_column_min_width(header) for header in headers]
    max_widths = [_column_max_width(header, total_width) for header in headers]
    min_width_sum = sum(min_widths)

    if min_width_sum >= total_width:
        scale = total_width / max(min_width_sum, 1)
        col_widths = [max(90, int(width * scale)) for width in min_widths]
    else:
        extra_width = total_width - min_width_sum
        extra_weights = [max(weight - 1.0, 0.2) for weight in weights]
        extra_weight_sum = sum(extra_weights) or float(col_count)
        col_widths = [
            min_widths[index] + int(extra_width * (extra_weights[index] / extra_weight_sum))
            for index in range(col_count)
        ]

    overflow = 0
    for index, width in enumerate(col_widths):
        if width > max_widths[index]:
            overflow += width - max_widths[index]
            col_widths[index] = max_widths[index]

    if overflow > 0:
        expandable_indexes = [
            index for index, width in enumerate(col_widths) if width < max_widths[index]
        ]
        while overflow > 0 and expandable_indexes:
            distributed = False
            for index in expandable_indexes:
                if col_widths[index] >= max_widths[index]:
                    continue
                col_widths[index] += 1
                overflow -= 1
                distributed = True
                if overflow == 0:
                    break
            if not distributed:
                break
            expandable_indexes = [
                index for index, width in enumerate(col_widths) if width < max_widths[index]
            ]

    diff = total_width - sum(col_widths)
    anchor_index = max(range(col_count), key=lambda idx: weights[idx])
    col_widths[anchor_index] += diff
    return col_widths


def _preset_column_widths(headers: List[str], total_width: int) -> Optional[List[int]]:
    if [header.strip() for header in headers] != LEGACY_DOC_STANDARD_HEADERS:
        return None

    ratios = [0.05, 0.12, 0.09, 0.10, 0.06, 0.15, 0.13, 0.08, 0.22]
    widths = [max(80, int(total_width * ratio)) for ratio in ratios]
    widths[-1] += total_width - sum(widths)
    return widths


def _classify_column_text_density(header: str) -> str:
    normalized = (header or "").strip()
    if any(keyword in normalized for keyword in SHORT_TEXT_COLUMN_KEYWORDS):
        return "short"
    if any(keyword in normalized for keyword in LONG_TEXT_COLUMN_KEYWORDS):
        return "long"
    if any(keyword in normalized for keyword in MEDIUM_TEXT_COLUMN_KEYWORDS):
        return "medium"
    return "default"


def _column_weight_hint(header: str) -> float:
    category = _classify_column_text_density(header)
    if category == "long":
        return 2.8
    if category == "short":
        return 0.9
    if category == "medium":
        return 1.4
    return 1.0


def _column_min_width(header: str) -> int:
    category = _classify_column_text_density(header)
    if category == "long":
        return 220
    if category == "short":
        return 110
    if category == "medium":
        return 130
    return 120


def _column_max_width(header: str, total_width: int) -> int:
    category = _classify_column_text_density(header)
    if category == "long":
        return max(520, int(total_width * 0.5))
    if category == "short":
        return 210
    if category == "medium":
        return 250
    return 230


def _load_font(size: int) -> Any:
    if ImageFont is None:
        raise RuntimeError("Pillow is required to load fonts")
    candidates = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Medium.ttc",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
    ]
    for path in candidates:
        try:
            return ImageFont.truetype(path, size=size)
        except Exception:
            continue
    return ImageFont.load_default()


def _wrap_text(text: str, font: Any, max_width: int) -> List[str]:
    if Image is None or ImageDraw is None:
        return [text]
    if not text:
        return [""]

    image = Image.new("RGB", (10, 10))
    drawer = ImageDraw.Draw(image)
    wrapped_lines: List[str] = []
    for paragraph in text.splitlines() or [text]:
        paragraph = paragraph.strip()
        if not paragraph:
            wrapped_lines.append("")
            continue
        current = ""
        for char in paragraph:
            trial = f"{current}{char}"
            bbox = drawer.textbbox((0, 0), trial, font=font)
            if bbox[2] - bbox[0] <= max_width or not current:
                current = trial
            else:
                wrapped_lines.append(current)
                current = char
        if current:
            wrapped_lines.append(current)
    return wrapped_lines or [""]
